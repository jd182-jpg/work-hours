#!/usr/bin/env python3
"""
Build the live work-hours workbook.

Creates a .xlsx with two sheets:

  Summary    Current pay period at a glance, then a month-by-month grid that
             lists only months with logged hours -- a dynamic array formula
             spills in each one as it appears, with the 1st-15th subtotal,
             the 16th-EOM subtotal, and the month total. Everything here is a
             live formula driven off the Hours Log table, so it updates
             itself the moment a row lands; nothing about this sheet needs to
             be regenerated when a new month starts.

  Hours Log  A table named HoursLog holding one row per day worked:
             Date, Day, Time In, Time Out, Hours, Period, Notes.
             The web app appends rows here through the Microsoft Graph API.

Usage:
    python3 make_workbook.py --year 2026 --out "Work Hours.xlsx"
"""

import argparse
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableStyleInfo

# Leggett house palette: navy ground, blue accent.
NAVY = "0F2440"
BLUE = "2E6FB7"
LIGHT = "EAF1F9"
RULE = "C9D6E4"
GREY = "6B7A8C"

HEADERS = ["Date", "Day", "Time In", "Time Out", "Hours", "Period", "Notes"]

thin = Side(style="thin", color=RULE)


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(bottom=thin)


def build_summary(ws, year):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    for col in "BCD":
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["E"].width = 3

    ws["A1"] = "Jackson's Hours"
    ws["A1"].font = Font(bold=True, size=18, color=NAVY)
    ws.row_dimensions[1].height = 26

    # ---- Current pay period (row 2 left blank as a spacer) -------------------
    ws["A3"] = "CURRENT PAY PERIOD"
    ws["A3"].font = Font(bold=True, size=10, color=BLUE)

    ws["A4"] = "Period start"
    ws["B4"] = ("=IF(DAY(TODAY())<=15,DATE(YEAR(TODAY()),MONTH(TODAY()),1),"
                "DATE(YEAR(TODAY()),MONTH(TODAY()),16))")
    ws["A5"] = "Period end"
    ws["B5"] = ("=IF(DAY(TODAY())<=15,DATE(YEAR(TODAY()),MONTH(TODAY()),15),"
                "EOMONTH(TODAY(),0))")
    ws["A6"] = "Hours this period"
    ws["B6"] = ('=SUMIFS(HoursLog[Hours],HoursLog[Date],">="&$B$4,'
                'HoursLog[Date],"<="&$B$5)')
    ws["A7"] = "Days worked"
    ws["B7"] = ('=COUNTIFS(HoursLog[Date],">="&$B$4,HoursLog[Date],"<="&$B$5)')

    for row in range(4, 8):
        ws.cell(row=row, column=1).font = Font(size=11, color=NAVY)
        ws.cell(row=row, column=2).font = Font(bold=True, size=11, color=NAVY)
    ws["B4"].number_format = "mmm d, yyyy"
    ws["B5"].number_format = "mmm d, yyyy"
    ws["B6"].number_format = "0.00"
    ws["B6"].fill = PatternFill("solid", fgColor=LIGHT)
    ws["B7"].number_format = "0"

    # ---- Month-by-month grid --------------------------------------------------
    # Column A is a single dynamic-array formula: it spills one row per
    # distinct month that HoursLog actually has entries for, sorted
    # chronologically. A month you haven't logged anything for yet simply
    # never appears -- there is nothing to regenerate or re-push when a new
    # one starts, the spill just grows on its own the next time this opens.
    header_row = 10
    ws.cell(row=header_row - 1, column=1, value="%s BY PAY PERIOD" % year).font = Font(
        bold=True, size=10, color=BLUE)

    for idx, label in enumerate(["Month", "1st - 15th", "16th - End", "Month Total"]):
        style_header(ws.cell(row=header_row, column=1 + idx, value=label))

    first_row = header_row + 1
    last_row = first_row + 11

    # Column A is one classic Ctrl+Shift+Enter array formula spanning all 12
    # reserved rows, not a "spilling" dynamic array. Tested directly against
    # real Excel: SORT/FILTER/UNIQUE compute correctly here (confirmed via
    # their _xlfn./_xlfn._xlws. compatibility-prefixed names, required
    # because a tool other than Excel has to spell out functions added after
    # the xlsx format's original function table was frozen), but this
    # environment's spill engine did not expand a dynamic array formula
    # authored outside Excel past its first row even after a full recalc.
    # The legacy array-formula mechanism has no such dependency -- each of
    # the 12 cells evaluates the identical formula and pulls its own element
    # back out with INDEX/ROW(), so nothing here depends on spill support at
    # all. A row with no corresponding month reads a blank "" and is skipped
    # by every dependent formula below it.
    #
    # First-of-month is DATE(YEAR(d),MONTH(d),1), not EOMONTH(d,-1)+1: also
    # tested directly, EOMONTH does not broadcast element-wise over a
    # table-column array in this position (#VALUE!), even nested inside
    # FILTER's own argument, while YEAR/MONTH/DATE do.
    months_expr = (
        '_xlfn._xlws.SORT(_xlfn.UNIQUE(_xlfn._xlws.FILTER('
        'DATE(YEAR(HoursLog[Date]),MONTH(HoursLog[Date]),1),'
        'YEAR(HoursLog[Date])=%d)))' % year)
    array_ref = "A%d:A%d" % (first_row, last_row)
    ws.cell(row=first_row, column=1).value = ArrayFormula(
        array_ref,
        '=IFERROR(INDEX(%s,ROW()-ROW($A$%d)+1),"")' % (months_expr, first_row))
    for r in range(first_row, last_row + 1):
        cell = ws.cell(row=r, column=1)
        cell.number_format = "mmmm yyyy"
        cell.font = Font(size=11, color=NAVY)

    # B/C/D are ordinary formulas in every reserved row, not part of the
    # spill. Each reads whatever landed in its own row's column A: real once
    # the array has spilled that far, genuinely blank otherwise, so the IF
    # guard keeps unused rows empty instead of showing stray zeroes.
    for i in range(12):
        r = first_row + i
        a = "A%d" % r
        ws.cell(row=r, column=2, value=(
            '=IF(%s="","",SUMIFS(HoursLog[Hours],HoursLog[Date],">="&%s,'
            'HoursLog[Date],"<="&(%s+14)))' % (a, a, a)))
        ws.cell(row=r, column=3, value=(
            '=IF(%s="","",SUMIFS(HoursLog[Hours],HoursLog[Date],">="&(%s+15),'
            'HoursLog[Date],"<="&EOMONTH(%s,0)))' % (a, a, a)))
        ws.cell(row=r, column=4, value='=IF(%s="","",B%d+C%d)' % (a, r, r))

        ws.cell(row=r, column=1).border = Border(bottom=thin)
        for c in range(2, 5):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(bottom=thin)
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(size=11, bold=(c == 4), color=NAVY if c == 4 else GREY)
        ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor=LIGHT)

    total_row = first_row + 12
    ws.cell(row=total_row, column=1, value="YEAR TOTAL").font = Font(
        bold=True, size=11, color="FFFFFF")
    for c in range(1, 5):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
    for c in range(2, 5):
        col = get_column_letter(c)
        # SUM ignores the "" that unused reserved rows return, so this stays
        # correct whether 1 month or all 12 have spilled in.
        cell = ws.cell(row=total_row, column=c,
                       value="=SUM(%s%d:%s%d)" % (col, first_row, col, first_row + 11))
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="right")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, size=11, color="FFFFFF")

    ws.freeze_panes = "A%d" % first_row


def build_log(ws):
    ws.sheet_view.showGridLines = False
    widths = [13, 8, 11, 11, 10, 12, 34]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.append(HEADERS)
    # A table needs at least one data row. The app fills this placeholder on the
    # first sync instead of appending, so no blank row is left behind.
    ws.append([None] * len(HEADERS))

    table = Table(displayName="HoursLog", ref="A1:G2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    for c in range(1, len(HEADERS) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True, color="FFFFFF")

    ws.cell(row=2, column=1).number_format = "mm/dd/yyyy"
    ws.cell(row=2, column=3).number_format = "h:mm AM/PM"
    ws.cell(row=2, column=4).number_format = "h:mm AM/PM"
    ws.cell(row=2, column=5).number_format = "0.00"

    # Force the Period column to Text. The app already avoids sending a value
    # Excel could mistake for a date, but a cell left on General format applies
    # that guesswork to anything landing in it later, by hand or otherwise.
    ws.cell(row=2, column=6).number_format = "@"
    ws.column_dimensions["F"].number_format = "@"

    ws.freeze_panes = "A2"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--year", type=int, default=2026)
    ap.add_argument("--out", default="Work Hours.xlsx")
    args = ap.parse_args()

    wb = Workbook()
    build_summary(wb.active, args.year)
    wb.active.title = "Summary"
    build_log(wb.create_sheet("Hours Log"))
    wb.save(args.out)
    print("wrote %s (%d bytes)" % (args.out, os.path.getsize(args.out)))


if __name__ == "__main__":
    main()
